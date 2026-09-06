#!/usr/bin/env python3
"""
Generate all figures, stat macros, and appendix tables for the
"Data Collection Findings and Synthesis" report.

Reads:
    ../data/questionnaire.csv   (374 questionnaire responses, 17 columns)
    ../data/interviews.xlsx     (14 coded in-person interviews)

Writes (into this directory):
    fig_profile.pdf
    fig_return_struggle.pdf
    fig_messy_agree.pdf
    fig_time_lost.pdf
    fig_why_not_automated.pdf
    fig_assistant_pref.pdf
    fig_workspace_trust.pdf
    fig_local_vs_cloud.pdf
    fig_wrong_suggestion.pdf
    fig_likely_use.pdf
    fig_stop_reason.pdf
    fig_features.pdf
    fig_open_feedback_themes.pdf
    fig_interview.pdf
    stats.tex                   (\newcommand macros for headline numbers)
    tables.tex                  (appendix distribution tables)
    ../data/interviews.csv      (cleaned copy of the interview coding)

Deterministic. Exits non-zero on any parse error.
"""

from __future__ import annotations

import csv
import re
import sys
from collections import Counter, OrderedDict
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
DATA = HERE.parent / "data"
CSV_PATH = DATA / "questionnaire.csv"
XLSX_PATH = DATA / "interviews.xlsx"

# --------------------------------------------------------------------------- #
# Style
# --------------------------------------------------------------------------- #

# Colourblind-safe, muted, print-friendly.
INK = "#1b1b1b"
GRID = "#c9c9c9"
PRIMARY = "#356fb0"      # blue
SECONDARY = "#b0632f"    # rust
NEUTRAL = "#7a7a7a"      # grey
ACCENT = "#4a8a5c"       # green
WARN = "#a23b3b"         # deep red

plt.rcParams.update(
    {
        "font.family": "serif",
        "font.serif": ["Times New Roman", "Nimbus Roman", "DejaVu Serif"],
        "font.size": 10,
        "axes.edgecolor": INK,
        "axes.labelcolor": INK,
        "axes.titlesize": 10,
        "axes.titleweight": "bold",
        "text.color": INK,
        "xtick.color": INK,
        "ytick.color": INK,
        "axes.grid": True,
        "grid.color": GRID,
        "grid.linewidth": 0.5,
        "figure.dpi": 150,
        "savefig.bbox": "tight",
        "savefig.pad_inches": 0.03,
    }
)

WIDTH = 6.3  # inches, close to \textwidth at a4 with 1in margins


def _finish(ax, xlabel=None, ylabel=None, title=None):
    if xlabel:
        ax.set_xlabel(xlabel)
    if ylabel:
        ax.set_ylabel(ylabel)
    if title:
        ax.set_title(title)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)


def _bar_counts(ax, labels, values, color, horizontal=False, annotate=True):
    if horizontal:
        bars = ax.barh(labels, values, color=color, height=0.62)
        ax.invert_yaxis()
        ax.grid(axis="y", visible=False)
        if annotate:
            for b, v in zip(bars, values):
                ax.text(
                    b.get_width() + max(values) * 0.01,
                    b.get_y() + b.get_height() / 2,
                    str(v),
                    va="center",
                    ha="left",
                    fontsize=9,
                )
    else:
        bars = ax.bar(labels, values, color=color, width=0.62)
        ax.grid(axis="x", visible=False)
        if annotate:
            for b, v in zip(bars, values):
                ax.text(
                    b.get_x() + b.get_width() / 2,
                    b.get_height() + max(values) * 0.02,
                    str(v),
                    va="bottom",
                    ha="center",
                    fontsize=9,
                )
    return bars


def save(fig, name):
    out = HERE / name
    fig.savefig(out)
    plt.close(fig)
    print(f"  wrote {name}")


# --------------------------------------------------------------------------- #
# Load questionnaire
# --------------------------------------------------------------------------- #

COLS = [
    "ts",
    "email",
    "freq_multifile",
    "role",
    "return_struggle",
    "messy_agree",
    "why_not_automated",
    "time_lost",
    "used_ai",
    "assistant_pref",
    "workspace_trust",
    "wrong_suggestion",
    "features",
    "local_vs_cloud",
    "likely_use",
    "stop_reason",
    "open_feedback",
]


def load_questionnaire():
    with open(CSV_PATH, encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh))
    data = [r for r in rows if r and re.match(r"\d+/\d+/\d{4}", r[0])]
    if not data:
        sys.exit("ERROR: no questionnaire data rows parsed")
    bad = [i for i, r in enumerate(data) if len(r) != 17]
    if bad:
        sys.exit(f"ERROR: {len(bad)} questionnaire rows do not have 17 columns")
    recs = [dict(zip(COLS, r)) for r in data]
    for rec in recs:
        for k in list(rec):
            rec[k] = rec[k].strip()
    return recs


def load_interviews():
    wb = load_workbook(XLSX_PATH)
    ws = wb["Sheet1"]
    rows = [
        [("" if c is None else str(c).strip()) for c in row]
        for row in ws.iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]
    header = ["lost_place", "workaround", "concept_reaction", "trust_break", "keep_using"]
    body = rows[1:]  # drop the sheet's own header row
    recs = [dict(zip(header, r)) for r in body]
    # cleaned csv copy
    with open(DATA / "interviews.csv", "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        for rec in recs:
            w.writerow([rec[h] for h in header])
    print(f"  wrote ../data/interviews.csv ({len(recs)} rows)")
    return recs


# --------------------------------------------------------------------------- #
# Helpers for ordered / normalised categories
# --------------------------------------------------------------------------- #

def ordered_counts(recs, key, order, normalise=None):
    raw = Counter(rec[key] for rec in recs)
    if normalise:
        merged = Counter()
        for k, v in raw.items():
            merged[normalise(k)] += v
        raw = merged
    return OrderedDict((label, raw.get(label, 0)) for label in order)


def pct(n, total):
    return round(100 * n / total)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    print("Loading data ...")
    q = load_questionnaire()
    iv = load_interviews()
    N = len(q)
    NIV = len(iv)
    print(f"  questionnaire rows: {N}")
    print(f"  interview rows: {NIV}")

    macros = OrderedDict()
    macros["Nresp"] = N
    macros["Niv"] = NIV

    tables = []  # (title, [(label, count)], total) for appendix

    # ---- helper to register an appendix table -------------------------------
    def reg_table(title, od, total=None):
        total = total if total is not None else sum(od.values())
        tables.append((title, list(od.items()), total))

    # ===================================================================== #
    # 1. Respondent profile  ->  fig_profile.pdf
    # ===================================================================== #
    role_norm = lambda s: {
        "employed": "Employed",
        "Vibe Coder": "Vibe coder",
        "Student also working as freelancer in agentic ai": "Freelance",
    }.get(s, s)
    role_order = [
        "Student",
        "Developer / Software Engineer",
        "Freelance",
        "Employed",
        "Vibe coder",
    ]
    roles = ordered_counts(q, "role", role_order, normalise=role_norm)
    reg_table("Q4. What best describes you?", roles)

    freq_order = ["Daily", "A Few Times a Week", "A Few Times a Month"]
    freq = ordered_counts(q, "freq_multifile", freq_order)
    reg_table("Q3. Frequency of multi-file project work", freq)

    fig, axes = plt.subplots(1, 2, figsize=(WIDTH, 2.7))
    _bar_counts(
        axes[0],
        [r.replace(" / Software Engineer", "").replace("Developer", "Developer")
         for r in roles],
        list(roles.values()),
        PRIMARY,
        horizontal=True,
    )
    _finish(axes[0], xlabel="Respondents", title="Role")
    _bar_counts(
        axes[1],
        ["Daily", "Few / week", "Few / month"],
        list(freq.values()),
        SECONDARY,
        horizontal=True,
    )
    _finish(axes[1], xlabel="Respondents", title="Multi-file work frequency")
    fig.tight_layout()
    save(fig, "fig_profile.pdf")

    macros["PctStudent"] = pct(roles["Student"], N)
    macros["PctDeveloper"] = pct(roles["Developer / Software Engineer"], N)
    macros["PctDailyWeekly"] = pct(freq["Daily"] + freq["A Few Times a Week"], N)
    macros["NDaily"] = freq["Daily"]

    # prior AI usage (prose + appendix only)
    ai_order = [
        "Yes, regularly",
        "Yes, occasionally",
        "Tried it once or twice",
        "Never",
    ]
    ai = ordered_counts(q, "used_ai", ai_order)
    reg_table("Q9. Prior use of an AI assistant for files / code / tasks", ai)
    macros["PctUsedAI"] = pct(ai["Yes, regularly"] + ai["Yes, occasionally"], N)
    macros["PctNeverAI"] = pct(ai["Never"], N)

    # ===================================================================== #
    # 2. Problem validation
    # ===================================================================== #
    rs_order = ["Never", "Rarely", "Sometimes", "Often", "Every time"]
    rs = ordered_counts(q, "return_struggle", rs_order)
    reg_table("Q5. Return after a break and struggle to remember what to do next", rs)
    fig, ax = plt.subplots(figsize=(WIDTH, 2.5))
    _bar_counts(ax, list(rs), list(rs.values()), PRIMARY)
    _finish(ax, ylabel="Respondents")
    save(fig, "fig_return_struggle.pdf")
    macros["PctReturnOften"] = pct(rs["Often"] + rs["Every time"], N)
    macros["PctReturnNever"] = pct(rs["Never"], N)

    # messy agree 1..5
    ma = ordered_counts(q, "messy_agree", ["1", "2", "3", "4", "5"])
    reg_table("Q6. Files become messy faster than I can clean them up (1 to 5)", ma)
    fig, ax = plt.subplots(figsize=(WIDTH, 2.5))
    colors = [NEUTRAL, NEUTRAL, NEUTRAL, PRIMARY, PRIMARY]
    _bar_counts(ax, list(ma), list(ma.values()), colors)
    _finish(ax, xlabel="1 = Strongly disagree      5 = Strongly agree", ylabel="Respondents")
    save(fig, "fig_messy_agree.pdf")
    ma_vals = {int(k): v for k, v in ma.items()}
    macros["PctMessyAgree"] = pct(ma_vals[4] + ma_vals[5], N)
    macros["MeanMessy"] = round(
        sum(k * v for k, v in ma_vals.items()) / N, 1
    )

    # time lost
    tl_order = [
        "Under 30 minutes",
        "30 minutes - 1hour",
        "1-3 hours",
        "3-5 hours",
        "Over 5 Hours",
    ]
    tl = ordered_counts(q, "time_lost", tl_order)
    reg_table("Q8. Estimated time lost per week to these issues", tl)
    fig, ax = plt.subplots(figsize=(WIDTH, 2.6))
    tl_labels = ["< 30 min", "30-60 min", "1-3 h", "3-5 h", "> 5 h"]
    _bar_counts(ax, tl_labels, list(tl.values()), SECONDARY)
    _finish(ax, ylabel="Respondents")
    save(fig, "fig_time_lost.pdf")
    macros["PctTimeOverHour"] = pct(tl["1-3 hours"] + tl["3-5 hours"] + tl["Over 5 Hours"], N)
    macros["PctTimeOverThree"] = pct(tl["3-5 hours"] + tl["Over 5 Hours"], N)

    # why not automated
    wna_norm = lambda s: {
        "5": None,
        "3": None,
    }.get(s, s)
    wna_raw = Counter(rec["why_not_automated"] for rec in q)
    wna_order = [
        "I don't trust automation to do it correctly without supervision",
        "I don't know how / it feels out of reach",
        "It would take longer to set up than the time it saves",
        "Not applicable ; I don't have repeated manual tasks",
    ]
    wna = OrderedDict((k, wna_raw.get(k, 0)) for k in wna_order)
    reg_table("Q7. Why repeated manual file tasks are not automated", wna)
    fig, ax = plt.subplots(figsize=(WIDTH, 2.7))
    wna_short = [
        "Don't trust automation\nwithout supervision",
        "Don't know how /\nfeels out of reach",
        "Setup would cost more\nthan it saves",
        "Not applicable /\nno repeated tasks",
    ]
    _bar_counts(ax, wna_short, list(wna.values()), [WARN, NEUTRAL, NEUTRAL, NEUTRAL],
                horizontal=True)
    _finish(ax, xlabel="Respondents")
    save(fig, "fig_why_not_automated.pdf")
    macros["PctNoTrustAuto"] = pct(wna[wna_order[0]], N)
    macros["PctDontKnowHow"] = pct(wna[wna_order[1]], N)
    macros["PctSetupCost"] = pct(wna[wna_order[2]], N)

    # ===================================================================== #
    # 3. Design and trust signals
    # ===================================================================== #
    # assistant preference (labels start with "Assistant A"/"Assistant B")
    pref = Counter()
    for rec in q:
        s = rec["assistant_pref"]
        if s.startswith("Assistant A"):
            pref["Assistant A\n(approve first)"] += 1
        elif s.startswith("Assistant B"):
            pref["Assistant B\n(act, then undo)"] += 1
    reg_table(
        "Q10. Preferred assistant: approve-first (A) vs act-then-undo (B)",
        OrderedDict(pref),
    )
    fig, ax = plt.subplots(figsize=(WIDTH, 2.4))
    _bar_counts(ax, list(pref), list(pref.values()), [PRIMARY, SECONDARY])
    _finish(ax, ylabel="Respondents")
    save(fig, "fig_assistant_pref.pdf")
    macros["NPrefA"] = pref["Assistant A\n(approve first)"]
    macros["NPrefB"] = pref["Assistant B\n(act, then undo)"]
    macros["PctPrefA"] = pct(pref["Assistant A\n(approve first)"], N)

    # workspace trust 1..5
    wt = ordered_counts(q, "workspace_trust", ["1", "2", "3", "4", "5"])
    reg_table(
        "Q11. A workspace-only assistant is more trustworthy than a full-access one (1 to 5)",
        wt,
    )
    fig, ax = plt.subplots(figsize=(WIDTH, 2.5))
    _bar_counts(ax, list(wt), list(wt.values()),
                [NEUTRAL, NEUTRAL, NEUTRAL, ACCENT, ACCENT])
    _finish(ax, xlabel="1 = Strongly disagree      5 = Strongly agree", ylabel="Respondents")
    save(fig, "fig_workspace_trust.pdf")
    wt_vals = {int(k): v for k, v in wt.items()}
    macros["PctWorkspaceTrust"] = pct(wt_vals[4] + wt_vals[5], N)
    macros["MeanWorkspaceTrust"] = round(sum(k * v for k, v in wt_vals.items()) / N, 1)

    # local vs cloud
    lvc_order = [
        "Local First even if its slower or weaker",
        "I would need to know more specifics before deciding either way",
        "Cloud-backed even if its less private",
    ]
    lvc = ordered_counts(q, "local_vs_cloud", lvc_order)
    reg_table("Q14. Local-first vs cloud-backed preference", lvc)
    fig, ax = plt.subplots(figsize=(WIDTH, 2.5))
    _bar_counts(
        ax,
        ["Local-first", "Need more\nspecifics", "Cloud-backed"],
        list(lvc.values()),
        [ACCENT, NEUTRAL, SECONDARY],
    )
    _finish(ax, ylabel="Respondents")
    save(fig, "fig_local_vs_cloud.pdf")
    macros["PctLocalFirst"] = pct(lvc[lvc_order[0]], N)
    macros["PctCloud"] = pct(lvc[lvc_order[2]], N)
    macros["PctNeedSpecifics"] = pct(lvc[lvc_order[1]], N)

    # wrong suggestion reaction
    ws_order = [
        "I'd be more cautious but keep using it",
        "I'd stop trusting its suggestions in that specific feature  area",
        "Yes, one mistake wouldn't change my usage",
        "I'd likely stop using the tool altogether",
    ]
    ws = ordered_counts(q, "wrong_suggestion", ws_order)
    reg_table("Q12. Reaction to one wrong or unwanted suggestion", ws)
    fig, ax = plt.subplots(figsize=(WIDTH, 2.7))
    ws_short = [
        "More cautious,\nbut keep using",
        "Stop trusting that\nfeature area",
        "One mistake would\nnot change usage",
        "Likely stop using\naltogether",
    ]
    _bar_counts(ax, ws_short, list(ws.values()),
                [ACCENT, NEUTRAL, ACCENT, WARN], horizontal=True)
    _finish(ax, xlabel="Respondents")
    save(fig, "fig_wrong_suggestion.pdf")
    macros["PctKeepAfterMistake"] = pct(ws[ws_order[0]] + ws[ws_order[2]], N)
    macros["PctQuitAfterMistake"] = pct(ws[ws_order[3]], N)

    # likely use 1..5
    lu = ordered_counts(q, "likely_use", ["1", "2", "3", "4", "5"])
    reg_table("Q15. Likelihood of sustained use if it worked as described (1 to 5)", lu)
    lu_vals = {int(k): v for k, v in lu.items()}
    mean_lu = sum(k * v for k, v in lu_vals.items()) / N
    fig, ax = plt.subplots(figsize=(WIDTH, 2.5))
    _bar_counts(ax, list(lu), list(lu.values()),
                [WARN, NEUTRAL, NEUTRAL, PRIMARY, PRIMARY])
    ymax = max(lu.values())
    ax.set_ylim(0, ymax * 1.22)
    ax.axvline(mean_lu - 1, color=INK, linestyle="--", linewidth=1)
    ax.annotate(
        f"mean {mean_lu:.1f}",
        xy=(mean_lu - 1, ymax * 1.12),
        xytext=(mean_lu - 1 + 0.15, ymax * 1.12),
        fontsize=9,
        ha="left",
        va="center",
    )
    _finish(ax, xlabel="1 = Very unlikely      5 = Very likely", ylabel="Respondents")
    save(fig, "fig_likely_use.pdf")
    macros["PctLikelyUse"] = pct(lu_vals[4] + lu_vals[5], N)
    macros["PctUnlikelyUse"] = pct(lu_vals[1] + lu_vals[2], N)
    macros["MeanLikelyUse"] = round(mean_lu, 1)

    # stop reason
    sr_order = [
        "I forget to use it/ it doesn't fit into my existing habit",
        "It makes a mistake I dont trust it to avoid again",
        "It doesn't do anything I couldn't already do myself easily",
        "Privacy or data-control concerns",
        "It asks for approval too often and slows me down",
        "I don't think I would stop using it",
    ]
    sr = ordered_counts(q, "stop_reason", sr_order)
    reg_table("Q16. Single most likely reason to stop using ARTEMIS after a week", sr)
    fig, ax = plt.subplots(figsize=(WIDTH, 3.0))
    sr_short = [
        "Does not fit my\nexisting habit",
        "Makes a mistake I\ncannot trust it to avoid",
        "Nothing I could not\nalready do myself",
        "Privacy or data-control\nconcerns",
        "Asks for approval too\noften, slows me down",
        "Would not stop\nusing it",
    ]
    _bar_counts(ax, sr_short, list(sr.values()),
                [SECONDARY, WARN, NEUTRAL, NEUTRAL, NEUTRAL, ACCENT], horizontal=True)
    _finish(ax, xlabel="Respondents")
    save(fig, "fig_stop_reason.pdf")
    macros["PctChurnHabit"] = pct(sr[sr_order[0]], N)
    macros["PctChurnMistake"] = pct(sr[sr_order[1]], N)
    macros["PctChurnRedundant"] = pct(sr[sr_order[2]], N)
    macros["PctChurnPrivacy"] = pct(sr[sr_order[3]], N)
    macros["PctChurnApproval"] = pct(sr[sr_order[4]], N)

    # ===================================================================== #
    # 4. Feature demand
    # ===================================================================== #
    FEATURE_CANON = [
        "Resuming interrupted work (session resume)",
        "Pulling together notes/citations from multiple documents",
        "Tidying up a messy folder",
        "Code assistant (explain errors, suggest fixes as a diff)",
        "Detecting and offering to automate a repeated task",
        "Drafting documents from existing material",
        "On-demand web search",
        "Drafting and sending emails (with approval)",
        "Speech input/output",
    ]
    fcount = Counter()
    for rec in q:
        raw = rec["features"]
        # rejoin the comma inside the code-assistant option
        raw = raw.replace(
            "Code assistant (explain errors, suggest fixes as a diff)",
            "Code assistant (explain errors; suggest fixes as a diff)",
        )
        for tok in raw.split(","):
            tok = tok.strip().replace(
                "Code assistant (explain errors; suggest fixes as a diff)",
                "Code assistant (explain errors, suggest fixes as a diff)",
            )
            if tok:
                fcount[tok] += 1
    # keep only canonical labels
    ordered_feats = sorted(
        [(f, fcount.get(f, 0)) for f in FEATURE_CANON],
        key=lambda kv: kv[1],
        reverse=True,
    )
    reg_table(
        "Q13. Features that would actually be opened in the first week (pick top 3)",
        OrderedDict(ordered_feats),
    )
    short_map = {
        "Resuming interrupted work (session resume)": "Session resume",
        "Pulling together notes/citations from multiple documents": "Multi-document synthesis",
        "Tidying up a messy folder": "Tidy a messy folder",
        "Code assistant (explain errors, suggest fixes as a diff)": "Code assistant (diff fixes)",
        "Detecting and offering to automate a repeated task": "Detect a repeated task",
        "Drafting documents from existing material": "Draft documents from material",
        "On-demand web search": "On-demand web search",
        "Drafting and sending emails (with approval)": "Draft and send emails",
        "Speech input/output": "Speech input / output",
    }
    fig, ax = plt.subplots(figsize=(WIDTH, 3.3))
    labels = [short_map[f] for f, _ in ordered_feats]
    values = [v for _, v in ordered_feats]
    top3 = {"Session resume", "Multi-document synthesis", "Tidy a messy folder"}
    bar_colors = [PRIMARY if l in top3 else NEUTRAL for l in labels]
    _bar_counts(ax, labels, values, bar_colors, horizontal=True)
    _finish(ax, xlabel="Times selected (multi-select, n = %d)" % N)
    save(fig, "fig_features.pdf")
    WORDS = ["One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten"]
    for i, (f, v) in enumerate(ordered_feats, start=1):
        macros[f"FeatRank{WORDS[i - 1]}Count"] = v
    macros["FeatTopName"] = "Resuming interrupted work"
    macros["FeatTopCount"] = ordered_feats[0][1]
    macros["FeatSpeechCount"] = fcount.get("Speech input/output", 0)

    # ---- open-ended feedback themes --------------------------------------
    THEMES = OrderedDict(
        [
            ("One-click universal undo", ["one-click undo", "undo for everything"]),
            ("Learn my folder structure over time",
             ["learn my folder structure", "learns from my work patterns",
              "feature of memory", "semantically organize"]),
            ("Already covered by Copilot / Claude Code",
             ["already use copilot", "existing coding agents",
              "why would i learn a new tool", "agent built in our ide",
              "agent built in ide"]),
            ("Concern over data sent to a server",
             ["how much of my data gets sent", "personal data is being sent",
              "sharing my data on a cloud", "data on a cloud"]),
            ("Auto-group files with similar names",
             ["grouping files with similar names", "combine all the related files",
              "collective folder", "combines all the related files"]),
            ("Merge duplicate versions of a file",
             ["merge duplicate versions"]),
            ("Track the latest version across copies",
             ["which version of a document is the latest",
              "which file has the latest", "tracking which version"]),
            ("Work across cloud drives (Drive, OneDrive)",
             ["work across cloud drives", "google drive, onedrive"]),
            ("I organise my own files and will not delegate it",
             ["very particular about organizing", "control freak",
              "i am a control freak", "wouldn't hand that off"]),
        ]
    )
    theme_counts = OrderedDict((k, 0) for k in THEMES)
    for rec in q:
        fb = rec["open_feedback"].lower().strip()
        if not fb or fb in ("-", "nil", "no", "none", "nan", "not really",
                            "nothing comes to mind right now.", "no nothing comes to mind",
                            "everything's covered.", "no nothing comes to mind."):
            continue
        for theme, needles in THEMES.items():
            if any(nd in fb for nd in needles):
                theme_counts[theme] += 1
    tc_sorted = sorted(theme_counts.items(), key=lambda kv: kv[1], reverse=True)
    reg_table("Q17. Recurring themes in open-ended feedback", OrderedDict(tc_sorted))
    fig, ax = plt.subplots(figsize=(WIDTH, 3.2))
    _bar_counts(
        ax,
        [k for k, _ in tc_sorted],
        [v for _, v in tc_sorted],
        [SECONDARY if k.startswith("Already covered") or k.startswith("Concern over")
         else PRIMARY for k, _ in tc_sorted],
        horizontal=True,
    )
    _finish(ax, xlabel="Mentions")
    save(fig, "fig_open_feedback_themes.pdf")
    for k, v in tc_sorted:
        pass
    macros["ThemeUndo"] = theme_counts["One-click universal undo"]
    macros["ThemeLearn"] = theme_counts["Learn my folder structure over time"]
    macros["ThemeRedundant"] = theme_counts["Already covered by Copilot / Claude Code"]
    macros["ThemeDataServer"] = theme_counts["Concern over data sent to a server"]

    # ===================================================================== #
    # 5. Interviews  ->  fig_interview.pdf
    # ===================================================================== #
    def code_concept(s):
        s = s.lower()
        if s.startswith("positive") or s == "good":
            return "Positive"
        if "semi" in s:
            return "Semi-positive"
        if "two" in s:
            return "Mixed"
        if "negative" in s:
            return "Negative"
        return "Other"

    def code_keep(s):
        s = s.lower()
        if s == "yes":
            return "Yes"
        if s.startswith("conditional") or "onboarding" in s or "current poc" in s \
           or "not at current" in s:
            return "Conditional"
        if "privacy" in s:
            return "Privacy concern"
        if s.startswith("no") or "wouldn't use" in s or "nope" in s:
            return "No"
        return "Other"

    concept = Counter(code_concept(r["concept_reaction"]) for r in iv)
    keep = Counter(code_keep(r["keep_using"]) for r in iv)

    concept_order = ["Positive", "Semi-positive", "Mixed", "Negative"]
    keep_order = ["Yes", "Conditional", "Privacy concern", "No"]
    concept_od = OrderedDict((k, concept.get(k, 0)) for k in concept_order)
    keep_od = OrderedDict((k, keep.get(k, 0)) for k in keep_order)
    reg_table("Interview: reaction to the concept", concept_od, total=NIV)
    reg_table("Interview: would you actually keep using it", keep_od, total=NIV)

    fig, axes = plt.subplots(1, 2, figsize=(WIDTH, 2.7))
    _bar_counts(axes[0], list(concept_od), list(concept_od.values()),
                [ACCENT, PRIMARY, NEUTRAL, WARN])
    _finish(axes[0], ylabel="Interviews", title="Reaction to the concept")
    axes[0].tick_params(axis="x", labelrotation=25)
    _bar_counts(axes[1], list(keep_od), list(keep_od.values()),
                [ACCENT, PRIMARY, SECONDARY, WARN])
    _finish(axes[1], title="Would keep using it")
    axes[1].tick_params(axis="x", labelrotation=25)
    fig.tight_layout()
    save(fig, "fig_interview.pdf")

    macros["IvPositive"] = concept.get("Positive", 0)
    macros["IvKeepYes"] = keep.get("Yes", 0)
    macros["IvKeepConditional"] = keep.get("Conditional", 0)

    # trust-break coding (prose only)
    def code_trust(s):
        s = s.lower()
        if "shrug" in s:
            return "Shrug it off"
        if "conditional" in s or "abnormal" in s:
            return "Conditional"
        if "stop" in s or "won't use" in s or "wont use" in s:
            return "Stop using"
        if "reinforcement" in s or "feedback" in s:
            return "Needs human feedback"
        return "Other"

    trust = Counter(code_trust(r["trust_break"]) for r in iv)
    reg_table("Interview: reaction to a trust-breaking moment",
              OrderedDict(sorted(trust.items(), key=lambda kv: kv[1], reverse=True)),
              total=NIV)
    macros["IvTrustShrug"] = trust.get("Shrug it off", 0)
    macros["IvTrustStop"] = trust.get("Stop using", 0)

    # ===================================================================== #
    # Segment cross-tabs (prose + appendix table, no figure)
    # ===================================================================== #
    def seg_students(rec):
        return role_norm(rec["role"]) == "Student"

    def seg_devs(rec):
        return rec["role"] == "Developer / Software Engineer"

    def seg_heavy_ai(rec):
        return rec["used_ai"] == "Yes, regularly"

    def seg_light_ai(rec):
        return rec["used_ai"] in ("Tried it once or twice", "Never")

    def seg_stats(pred):
        sub = [r for r in q if pred(r)]
        n = len(sub)
        if n == 0:
            return (0, 0.0, 0, 0)
        wt_mean = sum(int(r["workspace_trust"]) for r in sub) / n
        pref_a = sum(1 for r in sub if r["assistant_pref"].startswith("Assistant A"))
        local = sum(
            1 for r in sub if r["local_vs_cloud"].startswith("Local First")
        )
        return (n, round(wt_mean, 1), pct(pref_a, n), pct(local, n))

    segments = [
        ("All respondents", lambda r: True),
        ("Students", seg_students),
        ("Developers / engineers", seg_devs),
        ("Regular AI-assistant users", seg_heavy_ai),
        ("Light or non AI-assistant users", seg_light_ai),
    ]
    seg_rows = [(name, *seg_stats(pred)) for name, pred in segments]

    seg_lines = ["% Auto-generated by figures/gen_figures.py. Do not edit by hand.", ""]
    seg_lines.append("\\begin{table}[H]")
    seg_lines.append("\\centering")
    seg_lines.append("\\caption{Selected measures by respondent segment.}")
    seg_lines.append("{\\small")
    seg_lines.append("\\begin{tabular}{@{}lrccc@{}}")
    seg_lines.append("\\toprule")
    seg_lines.append(
        "Segment & $n$ & Workspace-trust & Prefer & Prefer \\\\"
    )
    seg_lines.append(
        " & & (mean, 1--5) & approval-first & local-first \\\\"
    )
    seg_lines.append("\\midrule")
    for name, n, wtm, pa, loc in seg_rows:
        seg_lines.append(f"{name} & {n} & {wtm} & {pa}\\% & {loc}\\% \\\\")
    seg_lines.append("\\bottomrule")
    seg_lines.append("\\end{tabular}}")
    seg_lines.append("\\end{table}")
    (HERE / "segments.tex").write_text("\n".join(seg_lines) + "\n", encoding="utf-8")
    print(f"  wrote segments.tex ({len(seg_rows)} rows)")

    print("\n=== SEGMENT CROSS-TABS ===")
    for name, n, wtm, pa, loc in seg_rows:
        print(f"  {name:32s} n={n:3d}  wt_mean={wtm}  prefA={pa}%  local={loc}%")

    by_name = {r[0]: r for r in seg_rows}
    macros["PctPrefAHeavy"] = by_name["Regular AI-assistant users"][3]
    macros["PctPrefALight"] = by_name["Light or non AI-assistant users"][3]
    macros["PctLocalStudent"] = by_name["Students"][4]
    macros["PctLocalDev"] = by_name["Developers / engineers"][4]

    # ===================================================================== #
    # Write stats.tex
    # ===================================================================== #
    def fmt_macro(name, value):
        return f"\\newcommand{{\\{name}}}{{{value}}}"

    lines = ["% Auto-generated by figures/gen_figures.py. Do not edit by hand.", ""]
    for k, v in macros.items():
        lines.append(fmt_macro(k, v))
    (HERE / "stats.tex").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"  wrote stats.tex ({len(macros)} macros)")

    # ===================================================================== #
    # Write tables.tex (appendix distributions)
    # ===================================================================== #
    def esc(s):
        for a, b in [("&", "\\&"), ("%", "\\%"), ("_", "\\_"), ("#", "\\#"),
                     ("’", "'"), ("‘", "'"), ("–", "-"), ("—", "-")]:
            s = s.replace(a, b)
        return s

    tl_lines = ["% Auto-generated by figures/gen_figures.py. Do not edit by hand.", ""]
    for title, items, total in tables:
        tl_lines.append("\\begin{table}[H]")
        tl_lines.append("\\centering")
        tl_lines.append(f"\\caption{{{esc(title)}}}")
        tl_lines.append("\\begin{tabular}{@{}lrr@{}}")
        tl_lines.append("\\toprule")
        tl_lines.append("Response & Count & Percent \\\\")
        tl_lines.append("\\midrule")
        for label, count in items:
            p = f"{100 * count / total:.1f}" if total else "0.0"
            tl_lines.append(f"{esc(str(label))} & {count} & {p}\\% \\\\")
        tl_lines.append("\\midrule")
        tl_lines.append(f"Total & {total} & 100.0\\% \\\\")
        tl_lines.append("\\bottomrule")
        tl_lines.append("\\end{tabular}")
        tl_lines.append("\\end{table}")
        tl_lines.append("")
    (HERE / "tables.tex").write_text("\n".join(tl_lines) + "\n", encoding="utf-8")
    print(f"  wrote tables.tex ({len(tables)} tables)")

    # ===================================================================== #
    # Console summary for cross-checking prose
    # ===================================================================== #
    print("\n=== STAT SUMMARY (cross-check against prose) ===")
    for k, v in macros.items():
        print(f"  {k:24s} {v}")
    print("\n=== FEATURE RANKING ===")
    for i, (f, v) in enumerate(ordered_feats, start=1):
        print(f"  {i:2d}. {v:4d}  {f}")
    print("\n=== OPEN-FEEDBACK THEMES ===")
    for k, v in tc_sorted:
        print(f"  {v:3d}  {k}")
    print("\nDone.")


if __name__ == "__main__":
    main()
